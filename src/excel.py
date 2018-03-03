#!/usr/bin/python
# -*- coding: utf-8 -*-

# *************************************************************
# Filename @ excel.py
# Author @ zhengxiasong
# Create date @ 2017-08-01 17:52:35
# Description @
# *************************************************************
# Script starts from here


from openpyxl import Workbook
from openpyxl import load_workbook
from  openpyxl.reader.excel import load_workbook
from openpyxl.writer.excel import ExcelWriter
import time
from copy import deepcopy
from openpyxl.chart import PieChart, BarChart, Series, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart import (Series, BarChart3D, PieChart3D)


class Excel(object):
    def __init__(self, log_file):
        localtime = time.strftime("%Y-%m-%d-%H-%M-%S", time.localtime())
        self.__filename = "../excels/TestResult-%s.xlsx" % localtime
        self.__wb = Workbook()
        # sheet = self.__wb.active
        self.__ws = self.__wb.active
        self.__ws.title = "test result"
        self.__log_name = log_file
        # self.__wb.create_sheet("test result")
        self.__sheetnames = self.__wb.get_sheet_names()
        self.__ws = self.__wb.get_sheet_by_name(self.__sheetnames[0])
        # 写列标题
        first_row = ["OperationName", "TotalTimes", "SuccessTimes", "FailTimes", "SuccessRate", "FailRate", "Remarks"]
        index = 0
        for char in "ABCDEFG":
            self.__ws["%s1" % char].value = first_row[index]
            index += 1

    # 读取log文件
    def read_log(self):
        input = open(self.__log_name, "r")
        # 读log到字典中
        excel_dict = {}
        child_dict = {"Succ": 0, "Fail": 0}
        for line in input:
            line = line.split()
            if not line: continue
            if line[5] not in excel_dict:
                excel_dict[line[5]] = deepcopy(child_dict)
            if line[6] == "True]":
                excel_dict[line[5]]["Succ"] += 1
            elif line[6] == "False]":
                excel_dict[line[5]]["Fail"] += 1
        # print excel_dict
        dict_keys = excel_dict.keys()
        # 将字典中的数据写入表中
        for keys in dict_keys:
            self.write_result(keys, excel_dict[keys]["Succ"], excel_dict[keys]["Succ"] + excel_dict[keys]["Fail"])
        # sheet = self.__wb.get_sheet_by_name(self.__sheetnames[0])
        # +2代表在最后空1行，再写数据
        max_row = self.__ws.max_row + 2

        # 计算成功次数
        ssum = 0
        for index in range(2, max_row - 1):
            ssum += self.__ws["C%d" % index].value
        self.__ws["A%d" % (max_row)].value = "SuccessSum"
        self.__ws["B%d" % (max_row)].value = ssum

        # 计算错误次数
        tsum = 0
        for index in range(2, max_row - 1):
            tsum += self.__ws["D%d" % index].value
        self.__ws["A%d" % (max_row + 1)].value = "FailSum"
        self.__ws["B%d" % (max_row + 1)].value = tsum

        self.__ws["A%d" % (max_row + 2)].value = "TimesSum"
        self.__ws["B%d" % (max_row + 2)].value = tsum + ssum

        # 计算成功率和失败率
        self.__ws["D%d" % (max_row)].value = "TotalSuccRate"
        self.__ws["E%d" % (max_row)].value = "%4.2f%%" % (ssum * 1.0 / (tsum + ssum) * 100)
        # sheet["E%d" % (max_row)].value = (ssum * 1.0 /(tsum + ssum))
        self.__ws["D%d" % (max_row + 1)].value = "TotalFailRate"
        self.__ws["E%d" % (max_row + 1)].value = "%4.2f%%" % (tsum * 1.0 / (tsum + ssum) * 100)
        # sheet["E%d" % (max_row+1)].value = (tsum * 1.0 /(tsum + ssum))
        # 画柱状图
        self.make_bar_chart()
        # 画饼图
        self.make_pie_chart()

    def write_result(self, operation_name, success_times, total_times):
        # 根据参数计算出相关信息，组成列表，将列表写入excel
        # sheet = self.__wb.get_sheet_by_name(self.__sheetnames[0])
        # +1代表从最后一行的下一行写起
        max_row = self.__ws.max_row + 1
        # test_time = time.strftime("%H:%M:%S", time.localtime())
        # new_row = [test_time]
        new_row = []
        new_row.append(operation_name)
        new_row.append(total_times)
        new_row.append(success_times)
        new_row.append(total_times - success_times)
        new_row.append("%5.2f%%" % float(success_times * 1.0 / total_times * 100))
        new_row.append("%5.2f%%" % float((total_times - success_times) * 1.0 / total_times * 100))

        index = 0
        for char in "ABCDEF":
            self.__ws["%s%d" % (char, max_row)].value = new_row[index]
            # print new_row[index]
            index += 1

    def make_bar_chart(self):
        # self.__wb = load_workbook(self.__filename)
        sheetnames = self.__wb.get_sheet_names()
        # ws = self.__wb.get_sheet_by_name(sheetnames[0])
        ws_new = self.__wb.create_sheet("bar chart")
        max_r = self.__ws.max_row - 3

        data = Reference(self.__ws, min_col=3, min_row=1, max_col=4, max_row=max_r)
        titles = Reference(self.__ws, min_col=1, min_row=2, max_row=max_r)
        chart = BarChart3D()
        chart.title = "Test Result Bar Chart"
        chart.add_data(data=data, titles_from_data=True)
        chart.set_categories(titles)
        chart.x_axis.title = ' '

        ws_new.add_chart(chart, "A1")
        self.__wb.save(self.__filename)

    def make_pie_chart(self):
        sheetnames = self.__wb.get_sheet_names()
        # ws = self.__wb.get_sheet_by_name(sheetnames[0])
        ws_new = self.__wb.create_sheet("pie chart")
        # 最后一行
        max_r = self.__ws.max_row

        pie = PieChart()
        labels = Reference(self.__ws, min_col=1, min_row=max_r - 2, max_row=max_r - 1)
        data = Reference(self.__ws, min_col=2, min_row=max_r - 3, max_row=max_r - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Test Result Pie Chart"

        # Cut the first slice out of the pie
        slice = DataPoint(idx=0, explosion=20)
        pie.series[0].data_points = [slice]

        ws_new.add_chart(pie, "A1")
        self.__wb.save(self.__filename)

    def __del__(self):
        self.__wb.save(self.__filename)


def main():
    new_excel = Excel("./excels/test.log")
    new_excel.read_log()


if __name__ == '__main__':
    main()
